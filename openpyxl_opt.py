import io
import json
import logging
import re
import openpyxl
from openpyxl import load_workbook

import constant
import json
import urllib.parse
import boto3


print('Loading function')

s3 = boto3.client('s3')


def lambda_handler(event, context):
    # print("Received event: " + json.dumps(event, indent=2))

    # Get the object from the event and show its content type
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'], encoding='utf-8')
    try:
        # Set some display options for the dataframe

        response = s3.get_object(Bucket=bucket, Key=key)
        body = response['Body'].read()
        print(body)
        workbook = load_workbook(io.BytesIO(body), data_only=True)
        process(workbook,bucket,'outputs3')
        #get_sheet_names = workbook.sheetnames

        #print("CONTENT TYPE: " + response['ContentType'])
        return response['ContentType']
    except Exception as e:
        print(e)
        print(
            'Error getting object {} from bucket {}. Make sure they exist and your bucket is in the same region as this function.'.format(
                key, bucket))
        raise e


def process(excel_file,bucket,s3file):
    config_de=read_config()
    #excel_file = config_de["inputfile"]
    wb_obj_input=fetch_input(excel_file)
    #sheet_obj_input=active_sheet(wb_obj_input)
    parse_excel(wb_obj_input,config_de,bucket,s3file)

def fetch_input(excel_file):
    wb_obj = openpyxl.load_workbook(excel_file)
    return wb_obj

def active_sheet(wb_obj):
    return wb_obj.active


def parse_excel(wb_obj,config_de,bucket,s3file):
    sheet_obj=wb_obj.active
    wb_output=save_excel(config_de)
    #output_sheet=active_sheet(wb_output)
    maxCol=parse_header(wb_output,config_de)
    parse_payload(sheet_obj,wb_output,config_de,bucket,s3file)

def parse_header(wb_output,config_de):
    output_sheet = active_sheet(wb_output)
    payload=config_de["payload"]
    iColumnCounter=1
    for ele in payload:
        header_name=ele["name"]
        #print(header_name)
        output_sheet.cell(row=constant.HEADER_ROW, column=iColumnCounter).value=header_name
        #print(output_sheet.cell(row=constant.HEADER_ROW, column=iColumnCounter).value)

        iColumnCounter=iColumnCounter+1
    update_excel(config_de,wb_output)
    return iColumnCounter-1

def parse_payload(input_sheet,wb_output,config_de,bucket,s3file):
    output_sheet = active_sheet(wb_output)
    payload=config_de["payload"]
    totalRow=config_de["maxrow"]
    startRow=config_de["startrow"]
    offsetCounter=0
    skipValue=False
    irejectCounter=0
    isuccessCounter=2
    skiprow_list=[]
    #Fetching the list of rows to reject:
    if "skiprow" in config_de:
        startrow=config_de["skiprow"]["startrow"]
        freq=config_de["skiprow"]["frequency"]
        #print(startRow)
        iCounter=startrow
        while iCounter <= totalRow:
            skiprow_list.append(iCounter)
            iCounter=iCounter+freq
        irowRejectCounter=6
    for irowCounter in range(2,totalRow-startRow+3):
        #print(irowCounter)
        irowRejectCounter=irowRejectCounter+1
        if irowRejectCounter not in skiprow_list:
            #print(irowRejectCounter)
            msgrow=""
            rowshift=True
            iColumnCounter=0
            skipValue=False
            #print("this row start" +str(irowCounter))
            #print(json.dumps(payload))
            for ele in payload:
                iColumnCounter=iColumnCounter+1
                if "constant" in ele:
                    if "value" in ele:
                        valid_flag,msg=validation(ele,ele["value"])
                        if not valid_flag:
                        #log error
                            skipValue=True
                            #output_sheet.delete_rows(irowCounter,1)
                            irejectCounter=irejectCounter+1

                            print(msg+"Row :" + str(currentRow)+" col: " +str(iColumnCounter) + " Value :"+ele["value"] + " ele :" + str(ele))
                        elif not skipValue:
                            msgrow=msgrow+" Element Name :"+ str(ele["name"]) + " Element Value :"+str(ele["value"]) + ","
                            output_sheet.cell(row=isuccessCounter,column=iColumnCounter).value=ele["value"]


                elif not skipValue and "row" in ele and "col" in ele:
                        rowValue=ele["row"]
                        colValue=ele["col"]

                        valid_flag, msg = validation(ele,input_sheet.cell(row=rowValue,column=colValue).value)

                        if not valid_flag:
                            #throw error
                            skipValue = True
                            #output_sheet.delete_rows(irowCounter,1)
                            irejectCounter = irejectCounter + 1
                            print(msg+"Row :" + str(currentRow) +" col: " +str(iColumnCounter) + " Value :"+str(input_sheet.cell(row=rowValue,column=colValue).value)+ " ele :" + str(ele))
                        elif not skipValue:
                            msgrow = msgrow + " Element Name :" + str(ele["name"]) + " Element Value :" + str(input_sheet.cell(row=rowValue,column=colValue).value) + ","
                            output_sheet.cell(row=isuccessCounter, column=iColumnCounter).value = input_sheet.cell(row=rowValue,column=colValue).value
                elif not skipValue and "startrow" in ele and "offsetlimit" in ele and ele["increment"]=="continuous":
                    if rowshift:
                        rowshift=False
                        currentRow=startRow+offsetCounter
                        maxSectionRow = startRow + ele["offsetlimit"]

                        if currentRow == maxSectionRow:
                            startRow = currentRow + 1
                            currentRow=startRow
                            offsetCounter = 1
                        else:
                            offsetCounter = offsetCounter + 1
                    if currentRow == totalRow:
                        break
                    colValue=ele["col"]
                    #print(currentRow)
                    #print(input_sheet.cell(row=currentRow, column=colValue).value)

                    valid_flag, msg =validation(ele,input_sheet.cell(row=currentRow,column=colValue).value)
                    if not valid_flag:
                        #throw error
                        skipValue = True
                        irejectCounter = irejectCounter + 1
                        #output_sheet.delete_rows(irowCounter,1)
                        print(msg+"Row :" + str(currentRow) +" col: " +str(iColumnCounter)+ " Value :"+str(input_sheet.cell(row=currentRow,column=colValue).value)+ " ele :" + str(ele))
                    elif not skipValue:
                        msgrow = msgrow + " Element Name :" + str(ele["name"]) + " Element Value :" + str(input_sheet.cell(row=currentRow,column=colValue).value) + ","
                        output_sheet.cell(row=isuccessCounter, column=iColumnCounter).value = input_sheet.cell(row=currentRow,column=colValue).value
                elif not skipValue and "startrow" in ele and "offsetlimit" in ele and ele["increment"] == "step":
                    frequency=((irowCounter+3-ele["startrow"])/ele["offsetlimit"])
                    print(int(frequency))
                    currentRow=int(frequency)*ele["offsetlimit"]+ele["startrow"]
                    print(currentRow)
                    colValue=ele["col"]
                    #print(int(currentRow))
                    #print(input_sheet.cell(row=currentRow,column=colValue).value)
                    valid_flag, msg = validation(ele,input_sheet.cell(row=currentRow,column=colValue).value)
                    if not valid_flag:
                        #print error on stdout
                        skipValue = True
                        #output_sheet.delete_rows(irowCounter,1)
                        irejectCounter = irejectCounter + 1
                        print(msg +"Row :" + str(currentRow) +" col: " +str(iColumnCounter)+ " Value :" + str(input_sheet.cell(row=currentRow,column=colValue).value)+ " ele :" + str(ele))
                    elif not skipValue:
                        msgrow = msgrow + " Element Name :" + str(ele["name"])+ " Element Value :" + str(input_sheet.cell(row=currentRow,column=colValue).value) + ","
                        output_sheet.cell(row=isuccessCounter, column=iColumnCounter).value = input_sheet.cell(row=currentRow,column=colValue).value
            if not skipValue:
                isuccessCounter=isuccessCounter+1
           # print("Row : "+str(isuccessCounter) + " Data : "+ msgrow)
    update_excel(config_de,wb_output)
    s3.upload_file(wb_output,bucket,s3file)
    print ("Total Number of Rows Inserted :" + str(isuccessCounter-2))
    print ("Total Number of Rows Deleted :"+str(irejectCounter))
       # input_sheet.cell()
def save_excel(config_de):
    outputfile=config_de["outputfile"]
    wb = openpyxl.Workbook()
    wb.save(outputfile)
    return wb

def update_excel(config_de,wb_output):
    outputfile = config_de["outputfile"]
    wb_output.save(outputfile)

def read_config():
    file_obj=open(constant.CONFIG_FILE,'r')
    return json.load(file_obj)

def fetchType(typevalue):
    if typevalue=="str":
        return str
    elif typevalue=="float":
        return float
    else:
        return int

def validation(ele,value):
    returnValue=True
    msgError=""
    if "validation" in ele:
        validation_rules=ele["validation"]
        if 'type' in validation_rules:
            if value is None or isinstance(value,fetchType(validation_rules["type"])):
                returnValue=True

            else:
                #print("inside else")
                returnValue=False
                msgError = constant.VALIDATION_TYPEMISMATCH
        if returnValue and 'format' in validation_rules:
            format=validation_rules['format']
            if re.search(format,value) is not None:
                returnValue=True
            else:
                returnValue=False
                msgError = constant.VALIDATION_REGEX
        if returnValue and 'minlength' in validation_rules:
            minLength=validation_rules['minlength']
            if value is not None:
                field_len=len(value)
            else:
                field_len=0
            if field_len >= minLength:
                returnValue=True
            else:
                returnValue=False
                msgError=constant.VALIDATION_MINLENGTH
        if returnValue and 'maxlength' in validation_rules:
            maxLength=validation_rules['maxlength']
            if value is not None:
                field_len=len(value)
            else:
                field_len=0
            if field_len <= maxLength:
                returnValue=True
            else:
                returnValue=False
                msgError = constant.VALIDATION_MAXLENGTH

        if returnValue and 'minimum' in validation_rules and (isinstance(value,int) or isinstance(value,float)):
            threshold_min = validation_rules['minimum']
            if value >= threshold_min:
                returnValue=True
            else:
                returnValue=False
                msgError=constant.VALIDATION_MINIMUMNUM

        if returnValue and 'maximum' in validation_rules and (isinstance(value,int) or isinstance(value,float)):
            threshold_max = validation_rules['maximum']
            if value <= threshold_max:
                returnValue=True
            else:
                returnValue=False
                msgError=constant.VALIDATION_MAXIMUMNUM
        return returnValue,msgError
    else:
        return returnValue,msgError

#process()
